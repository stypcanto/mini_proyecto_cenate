#!/usr/bin/env python3
"""
========================================================================
normalizar_excel_107.py – Auto-corrector de Archivos Excel Formulario 107
========================================================================
Script para normalizar automáticamente archivos Excel de Formulario 107
antes de importarlos al sistema CENATE.

Funcionalidades:
- Valida estructura de 14 columnas
- Corrige nombres de cabeceras automáticamente
- Normaliza formatos de fecha
- Genera reporte de cambios realizados
- Crea archivo corregido listo para importar

Uso:
    python normalizar_excel_107.py archivo.xlsx
    python normalizar_excel_107.py carpeta_con_archivos/

Autor: Styp Canto Rondon
Versión: 1.15.0
========================================================================
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

# Mapeo de variaciones -> nombre estándar
COLUMN_MAPPINGS = {
    # REGISTRO
    "registro": "REGISTRO",
    "reg": "REGISTRO",
    "nro registro": "REGISTRO",

    # OPCIONES DE INGRESO
    "opciones de ingreso de llamada": "OPCIONES DE INGRESO DE LLAMADA",
    "opcion de ingreso": "OPCIONES DE INGRESO DE LLAMADA",
    "opciones ingreso": "OPCIONES DE INGRESO DE LLAMADA",
    "tipo de ingreso": "OPCIONES DE INGRESO DE LLAMADA",
    "ingreso": "OPCIONES DE INGRESO DE LLAMADA",

    # TELEFONO
    "telefono": "TELEFONO",
    "teléfono": "TELEFONO",
    "tel": "TELEFONO",
    "celular": "TELEFONO",
    "movil": "TELEFONO",
    "móvil": "TELEFONO",

    # TIPO DE DOCUMENTO
    "tipo de documento": "TIPO DE DOCUMENTO",
    "tipo documento": "TIPO DE DOCUMENTO",
    "tipo_documento": "TIPO DE DOCUMENTO",
    "tipodocumento": "TIPO DE DOCUMENTO",
    "tipo doc": "TIPO DE DOCUMENTO",
    "tipo_doc": "TIPO DE DOCUMENTO",
    "tip doc": "TIPO DE DOCUMENTO",

    # DNI
    "dni": "DNI",
    "numero de documento": "DNI",
    "numero documento": "DNI",
    "nro documento": "DNI",
    "nro doc": "DNI",
    "documento": "DNI",

    # APELLIDOS Y NOMBRES
    "apellidos y nombres": "APELLIDOS Y NOMBRES",
    "nombres y apellidos": "APELLIDOS Y NOMBRES",
    "nombre completo": "APELLIDOS Y NOMBRES",
    "paciente": "APELLIDOS Y NOMBRES",
    "nombres": "APELLIDOS Y NOMBRES",
    "apellidos": "APELLIDOS Y NOMBRES",

    # SEXO
    "sexo": "SEXO",
    "genero": "SEXO",
    "género": "SEXO",
    "sex": "SEXO",

    # FECHA NACIMIENTO
    "fechanacimiento": "FechaNacimiento",
    "fecha nacimiento": "FechaNacimiento",
    "fecha de nacimiento": "FechaNacimiento",
    "fec nacimiento": "FechaNacimiento",
    "fec nac": "FechaNacimiento",
    "fecha nac": "FechaNacimiento",
    "f nacimiento": "FechaNacimiento",
    "f nac": "FechaNacimiento",
    "fecha_nacimiento": "FechaNacimiento",

    # DEPARTAMENTO
    "departamento": "DEPARTAMENTO",
    "depto": "DEPARTAMENTO",
    "dpto": "DEPARTAMENTO",
    "dep": "DEPARTAMENTO",

    # PROVINCIA
    "provincia": "PROVINCIA",
    "prov": "PROVINCIA",

    # DISTRITO
    "distrito": "DISTRITO",
    "dist": "DISTRITO",

    # MOTIVO DE LA LLAMADA
    "motivo de la llamada": "MOTIVO DE LA LLAMADA",
    "motivo llamada": "MOTIVO DE LA LLAMADA",
    "motivo": "MOTIVO DE LA LLAMADA",
    "motivo de consulta": "MOTIVO DE LA LLAMADA",

    # AFILIACION
    "afiliacion": "AFILIACION",
    "afiliación": "AFILIACION",
    "tipo afiliacion": "AFILIACION",
    "tipo afiliación": "AFILIACION",

    # DERIVACION INTERNA
    "derivacion interna": "DERIVACION INTERNA",
    "derivación interna": "DERIVACION INTERNA",
    "derivacion": "DERIVACION INTERNA",
    "derivación": "DERIVACION INTERNA",
    "deriva": "DERIVACION INTERNA",
}

EXPECTED_COLUMNS = [
    "REGISTRO",
    "OPCIONES DE INGRESO DE LLAMADA",
    "TELEFONO",
    "TIPO DE DOCUMENTO",
    "DNI",
    "APELLIDOS Y NOMBRES",
    "SEXO",
    "FechaNacimiento",
    "DEPARTAMENTO",
    "PROVINCIA",
    "DISTRITO",
    "MOTIVO DE LA LLAMADA",
    "AFILIACION",
    "DERIVACION INTERNA"
]


def normalize_header(raw_header):
    """Normaliza una cabecera a su forma estándar"""
    if not raw_header or not isinstance(raw_header, str):
        return None

    # Limpiar y convertir a minúsculas
    cleaned = raw_header.strip().lower()
    cleaned = ' '.join(cleaned.split())  # Quitar espacios extras

    return COLUMN_MAPPINGS.get(cleaned)


def normalize_excel(input_file, output_dir=None):
    """
    Normaliza un archivo Excel de Formulario 107

    Args:
        input_file: Ruta al archivo Excel original
        output_dir: Directorio de salida (opcional)

    Returns:
        dict: Reporte de normalización
    """
    print(f"\n{'='*70}")
    print(f"🔧 Normalizando archivo: {Path(input_file).name}")
    print(f"{'='*70}\n")

    # Cargar archivo
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        return {"error": f"No se pudo abrir el archivo: {str(e)}"}

    # Leer cabeceras originales
    original_headers = []
    for cell in ws[1]:
        original_headers.append(cell.value if cell.value else "")

    # Validar cantidad de columnas
    if len(original_headers) != 14:
        return {
            "error": f"Cantidad de columnas incorrecta: {len(original_headers)} (esperadas: 14)",
            "headers_found": original_headers
        }

    # Normalizar cabeceras
    normalized_headers = []
    changes = []
    errors = []

    for i, orig in enumerate(original_headers, 1):
        norm = normalize_header(orig)

        if norm is None:
            errors.append(f"Posición {i}: '{orig}' no reconocida")
            normalized_headers.append(None)
        else:
            normalized_headers.append(norm)
            if orig != norm:
                changes.append(f"  📝 Posición {i}: '{orig}' → '{norm}'")

    # Verificar errores
    if errors:
        print("❌ Errores encontrados:")
        for err in errors:
            print(f"  {err}")
        return {"error": "Columnas no reconocidas", "details": errors}

    # Verificar orden
    order_errors = []
    for i, (expected, actual) in enumerate(zip(EXPECTED_COLUMNS, normalized_headers), 1):
        if expected != actual:
            order_errors.append(f"Posición {i}: esperado '{expected}', encontrado '{actual}'")

    if order_errors:
        print("❌ Orden de columnas incorrecto:")
        for err in order_errors:
            print(f"  {err}")
        return {"error": "Orden incorrecto", "details": order_errors}

    # Aplicar correcciones
    if changes:
        print("✅ Cambios detectados:")
        for change in changes:
            print(change)

        # Actualizar cabeceras en el Excel
        for i, norm in enumerate(normalized_headers, 1):
            cell = ws.cell(row=1, column=i)
            cell.value = norm
            # Resaltar celdas modificadas
            if original_headers[i-1] != norm:
                cell.font = Font(bold=True, color="0000FF")
                cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    else:
        print("✅ No se necesitan correcciones (cabeceras correctas)")

    # Guardar archivo normalizado
    if output_dir is None:
        output_dir = Path(input_file).parent / "normalizados"

    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    original_name = Path(input_file).stem
    output_file = output_dir / f"{original_name}_normalizado_{timestamp}.xlsx"

    wb.save(output_file)
    wb.close()

    print(f"\n📁 Archivo guardado: {output_file}")
    print(f"{'='*70}\n")

    return {
        "success": True,
        "input_file": str(input_file),
        "output_file": str(output_file),
        "total_columns": len(original_headers),
        "changes_made": len(changes),
        "original_headers": original_headers,
        "normalized_headers": normalized_headers,
        "changes": changes
    }


def process_directory(directory):
    """Procesa todos los archivos .xlsx en un directorio"""
    dir_path = Path(directory)

    if not dir_path.is_dir():
        print(f"❌ Error: {directory} no es un directorio válido")
        return

    excel_files = list(dir_path.glob("*.xlsx"))

    if not excel_files:
        print(f"⚠️  No se encontraron archivos .xlsx en {directory}")
        return

    print(f"\n📂 Procesando {len(excel_files)} archivos...\n")

    results = []
    for file in excel_files:
        result = normalize_excel(file)
        results.append(result)

    # Resumen
    print(f"\n{'='*70}")
    print("📊 RESUMEN DE PROCESAMIENTO")
    print(f"{'='*70}")

    successful = sum(1 for r in results if r.get("success"))
    failed = len(results) - successful

    print(f"✅ Exitosos: {successful}")
    print(f"❌ Fallidos: {failed}")

    if failed > 0:
        print(f"\nArchivos con errores:")
        for r in results:
            if not r.get("success"):
                print(f"  - {Path(r.get('input_file', 'desconocido')).name}: {r.get('error')}")


def main():
    """Función principal"""
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUso:")
        print(f"  {sys.argv[0]} <archivo.xlsx>")
        print(f"  {sys.argv[0]} <directorio>")
        sys.exit(1)

    input_path = sys.argv[1]

    if not os.path.exists(input_path):
        print(f"❌ Error: {input_path} no existe")
        sys.exit(1)

    if os.path.isdir(input_path):
        process_directory(input_path)
    elif input_path.endswith('.xlsx'):
        result = normalize_excel(input_path)
        if not result.get("success"):
            print(f"\n❌ Error: {result.get('error')}")
            if result.get("details"):
                for detail in result.get("details"):
                    print(f"  {detail}")
            sys.exit(1)
    else:
        print(f"❌ Error: {input_path} debe ser un archivo .xlsx o un directorio")
        sys.exit(1)


if __name__ == "__main__":
    main()
